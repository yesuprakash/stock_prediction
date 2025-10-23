import yfinance as yf
import pandas as pd
import requests
import json
from db import get_connection
from datetime import datetime, timedelta
from ta.momentum import RSIIndicator
from ta.trend import MACD
from ta.volatility import BollingerBands, AverageTrueRange
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# -------------------------------
# 0️⃣ API Keys
# -------------------------------
NEWSAPI_KEY = "31a4ab26b3ca4edb9edd5b5e5bc272ef"

# -------------------------------
# 1️⃣ Stock List (NSE Tickers)
# -------------------------------
stocks = ["ABB.NS", "ACC.NS", "TATACONSUM.NS"]

# -------------------------------
# 2️⃣ Excel Columns
# RSI: Relative Strength Index, is a momentum-based technical indicator that measures the speed and change of a stock's price movements to help identify whether it is overbought or oversold.
# MACD stands for Moving Average Convergence Divergence, a technical analysis tool used to identify trends and momentum in the stock market. It is calculated by subtracting a 26-period Exponential Moving Average (EMA) from a 12-period EMA, and is typically used alongside a 9-period EMA of the MACD line (the signal line) to generate buy and sell signals.
# -------------------------------
columns = [
    "Stock Name", "Date", "Sector / Industry Outlook", "Trend", "Recent High/Low",
    "Current Price vs Moving Averages", "RSI", "MACD Trend",
    "Average Daily Volume", "Recent Volume Spikes", "Liquidity",
    "ATR", "Expected Price Range", "Volatility Level",
    "Key Support Levels", "Key Resistance Levels", "Probability of Trade Success (%)",
    "Moving Averages", "RSI Value", "MACD Signal", "Bollinger Band Position", "Bollinger % Position",
    "Chart Pattern Observed", "Trade Signal",
    "Upcoming Earnings/Dividends/Corporate Actions", "Catalyst Events",
    "Market Sentiment / Analyst Notes", "Best-case Price Target", "Likely Price Range",
    "Worst-case / Stop-Loss Risk", "Risk/Reward Ratio",
    "Technical Strength Score (%)", "Suggested Entry Price Range", "Stop-Loss Price", "Target Price",
    "Expected Holding Duration", "Additional Notes"
]

df = pd.DataFrame(columns=columns)

def insert_prediction(row):
    conn = get_connection()
    cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO predictions
            (prediction_date, stock_symbol, trade_signal, probability_success, 
             technical_strength, risk_reward, entry_price, target_price, 
             stop_loss, sector_outlook, sentiment, trend, raw_data)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            row["Date"],
            row["Stock Name"],
            row["Trade Signal"],
            float(row["Probability of Trade Success (%)"]) if row["Probability of Trade Success (%)"] else None,
            float(row["Technical Strength Score (%)"]) if row["Technical Strength Score (%)"] else None,
            float(row["Risk/Reward Ratio"]) if row["Risk/Reward Ratio"] else None,
            float(row["Suggested Entry Price Range"]) if row["Suggested Entry Price Range"] else None,
            float(row["Target Price"]) if row["Target Price"] else None,
            float(row["Stop-Loss Price"]) if row["Stop-Loss Price"] else None,
            row["Sector / Industry Outlook"],
            row["Market Sentiment / Analyst Notes"],
            row["Trend"],
            json.dumps(row.to_dict())  # store the full row as JSON for future flexibility
        ))
        conn.commit()
    except Exception as e:
        print(f"❌ DB Insert Error for {row['Stock Name']}: {e}")
    finally:
        cursor.close()
        conn.close()


# -------------------------------
# 3️⃣ Helper Functions (Technical)
# -------------------------------
def calculate_rsi(data, period=14):
    rsi = RSIIndicator(data['Close'], window=period).rsi()
    return rsi.iloc[-1]

def calculate_macd_trend(data):
    macd = MACD(data['Close']).macd()
    signal = MACD(data['Close']).macd_signal()
    return "Bullish" if macd.iloc[-1] > signal.iloc[-1] else "Bearish"

def calculate_bollinger_position(data):
    bb = BollingerBands(data['Close'], window=20, window_dev=2)
    current = data['Close'].iloc[-1]
    lower = bb.bollinger_lband().iloc[-1]
    upper = bb.bollinger_hband().iloc[-1]
    if current > upper:
        return "Above Upper Band", 100
    elif current < lower:
        return "Below Lower Band", 0
    else:
        perc = (current - lower) / (upper - lower) * 100
        return "Within Bands", round(perc, 2)

def calculate_atr(data):
    atr = AverageTrueRange(data['High'], data['Low'], data['Close'], window=14).average_true_range()
    return atr.iloc[-1]

def identify_chart_pattern(data):
    highs = data['High'].tail(20)
    lows = data['Low'].tail(20)
    if highs.max() == data['High'].iloc[-1]:
        return "Potential Uptrend / Breakout"
    elif lows.min() == data['Low'].iloc[-1]:
        return "Potential Downtrend / Breakdown"
    else:
        return "Sideways / Consolidation"

def volume_spike(data):
    avg_vol = data['Volume'].rolling(10).mean().iloc[-1]
    recent = data['Volume'].iloc[-1]
    return "Spike" if recent > 1.5 * avg_vol else "Normal"

def liquidity_level(data):
    avg_vol = data['Volume'].rolling(10).mean().iloc[-1]
    if avg_vol > 500000:
        return "High"
    elif avg_vol > 100000:
        return "Medium"
    else:
        return "Low"

def probability_breakout(rsi, macd_trend, bb_pos):
    score = 50
    if rsi < 70 and macd_trend == "Bullish" and "Upper" not in bb_pos:
        score += 20
    elif rsi > 70 or macd_trend == "Bearish":
        score -= 20
    return max(min(score, 100), 0)

def trade_signal(macd_trend, rsi, bb_perc):
    if macd_trend == "Bullish" and rsi < 70 and bb_perc < 80:
        return "Strong Buy"
    elif macd_trend == "Bearish" and rsi > 30 and bb_perc > 20:
        return "Strong Sell"
    else:
        return "Neutral"

# -------------------------------
# 4️⃣ Helper Functions (Indian Sector / News)
# -------------------------------
def get_sector_trend_dynamic(stock_symbol):
    sector_index_map = {
        "GAIL.NS": "CNXENERGY",
        "TATASTEEL.NS": "CNXMETAL",
        "HINDALCO.NS": "CNXMETAL"
    }
    sector_index = sector_index_map.get(stock_symbol, None)
    if not sector_index:
        return "Unknown"
    return f"{sector_index.replace('CNX','')} sector trend; short-term outlook based on recent momentum"

def get_upcoming_earnings(stock_symbol):
    return "No upcoming earnings"

def get_recent_news(company_name):
    today = datetime.now()
    from_date = (today - timedelta(days=14)).strftime('%Y-%m-%d')
    url = f"https://newsapi.org/v2/everything?q={company_name}+India&from={from_date}&sortBy=publishedAt&language=en&apiKey={NEWSAPI_KEY}"
    response = requests.get(url)
    if response.status_code == 200:
        articles = response.json().get("articles", [])
        if articles:
            headlines = [a["title"] for a in articles[:5]]
            return headlines
    return []

def analyze_sentiment(headlines):
    positive = ["gain", "rise", "bullish", "up", "growth", "profit"]
    negative = ["fall", "drop", "bearish", "loss", "decline", "weak"]
    score = 0
    for h in headlines:
        h_lower = h.lower()
        score += sum(1 for word in positive if word in h_lower)
        score -= sum(1 for word in negative if word in h_lower)
    if score > 1:
        return "Positive"
    elif score < -1:
        return "Negative"
    else:
        return "Neutral"

# -------------------------------
# 5️⃣ Process Each Stock
# -------------------------------
for stock in stocks:
    ticker = yf.Ticker(stock)
    end_date = datetime.now()
    start_date = end_date - timedelta(days=60)
    data = ticker.history(start=start_date.strftime('%Y-%m-%d'), end=end_date.strftime('%Y-%m-%d'))
    if data.empty:
        print(f"No data for {stock}")
        continue

    current_price = data['Close'].iloc[-1]
    recent_high = data['High'].max()
    recent_low = data['Low'].min()
    
    ma5 = data['Close'].rolling(5).mean().iloc[-1]
    ma10 = data['Close'].rolling(10).mean().iloc[-1]
    ma20 = data['Close'].rolling(20).mean().iloc[-1]
    
    rsi = calculate_rsi(data)
    macd_trend = calculate_macd_trend(data)
    bb_position, bb_percent = calculate_bollinger_position(data)
    atr = calculate_atr(data)
    chart_pattern = identify_chart_pattern(data)
    vol_spike = volume_spike(data)
    liquidity = liquidity_level(data)
    
    prob_breakout = probability_breakout(rsi, macd_trend, bb_position)
    signal = trade_signal(macd_trend, rsi, bb_percent)

    sector = get_sector_trend_dynamic(stock)
    earnings_str = get_upcoming_earnings(stock)
    headlines = get_recent_news(stock.split(".")[0])
    sentiment = analyze_sentiment(headlines)
    catalyst_events = "; ".join(headlines) if headlines else "No recent events"

    tech_score = 0
    if macd_trend == "Bullish": tech_score += 3
    if 30 < rsi < 70: tech_score += 2
    if bb_position == "Within Bands": tech_score += 1
    if vol_spike == "Spike": tech_score += 2
    if liquidity == "High": tech_score += 2
    if sentiment == "Positive": tech_score +=1
    tech_score_percent = round((tech_score / 11) * 100, 2)

    support = recent_low
    resistance = recent_high
    entry_price = support + (resistance - support) * 0.2
    target_price = resistance
    stop_loss = support * 0.98
    risk_reward = round((target_price - entry_price) / (entry_price - stop_loss), 2)

    new_row = pd.DataFrame([{
        "Stock Name": stock,
        "Date": end_date.strftime("%Y-%m-%d"),
        "Sector / Industry Outlook": sector,
        "Trend": macd_trend,
        "Recent High/Low": f"{recent_high}/{recent_low}",
        "Current Price vs Moving Averages": f"{current_price} vs MA5:{ma5:.2f}, MA10:{ma10:.2f}, MA20:{ma20:.2f}",
        "RSI": rsi,
        "MACD Trend": macd_trend,
        "Average Daily Volume": data['Volume'].mean(),
        "Recent Volume Spikes": vol_spike,
        "Liquidity": liquidity,
        "ATR": atr,
        "Expected Price Range": f"{recent_low} - {recent_high}",
        "Volatility Level": "High" if atr > (recent_high-recent_low)/2 else "Moderate",
        "Key Support Levels": support,
        "Key Resistance Levels": resistance,
        "Probability of Trade Success (%)": prob_breakout,
        "Moving Averages": f"MA5:{ma5:.2f}, MA10:{ma10:.2f}, MA20:{ma20:.2f}",
        "RSI Value": rsi,
        "MACD Signal": macd_trend,
        "Bollinger Band Position": bb_position,
        "Bollinger % Position": bb_percent,
        "Chart Pattern Observed": chart_pattern,
        "Trade Signal": signal,
        "Upcoming Earnings/Dividends/Corporate Actions": earnings_str,
        "Catalyst Events": catalyst_events,
        "Market Sentiment / Analyst Notes": sentiment,
        "Best-case Price Target": round(target_price,2),
        "Likely Price Range": f"{entry_price} - {target_price}",
        "Worst-case / Stop-Loss Risk": round(stop_loss,2),
        "Risk/Reward Ratio": risk_reward,
        "Technical Strength Score (%)": tech_score_percent,
        "Suggested Entry Price Range": round(entry_price,2),
        "Stop-Loss Price": round(stop_loss,2),
        "Target Price": round(target_price,2),
        "Expected Holding Duration": "1-3 weeks",
        "Additional Notes": ""
    }])

    df = pd.concat([df, new_row], ignore_index=True)

    insert_prediction(new_row.iloc[0])
    
# -------------------------------
# 6️⃣ Save to Excel
# -------------------------------
output_file = "short_term_analysis.xlsx"
df.to_excel(output_file, index=False)

# -------------------------------
# 7️⃣ Apply Dashboard Formatting
# -------------------------------
wb = openpyxl.load_workbook(output_file)
ws = wb.active

# Trade Signal coloring (Strong Buy=Green, Strong Sell=Red, Neutral=Yellow)
trade_signal_col = 24  # Column X
for row in range(2, ws.max_row+1):
    cell = ws.cell(row=row, column=trade_signal_col)
    if cell.value == "Strong Buy":
        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    elif cell.value == "Strong Sell":
        cell.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
    elif cell.value == "Neutral":
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

# Probability of Trade Success gradient
prob_col = 17  # Column Q
ws.conditional_formatting.add(f"{openpyxl.utils.get_column_letter(prob_col)}2:{openpyxl.utils.get_column_letter(prob_col)}{ws.max_row}",
    ColorScaleRule(start_type='min', start_color='FF7F7F',
                   mid_type='percentile', mid_value=50, mid_color='FFFF99',
                   end_type='max', end_color='90EE90'))

# Technical Strength Score gradient
tech_col = 33  # Column AG
ws.conditional_formatting.add(f"{openpyxl.utils.get_column_letter(tech_col)}2:{openpyxl.utils.get_column_letter(tech_col)}{ws.max_row}",
    ColorScaleRule(start_type='min', start_color='FF7F7F',
                   mid_type='percentile', mid_value=50, mid_color='FFFF99',
                   end_type='max', end_color='90EE90'))

# Risk/Reward Ratio coloring
rr_col = 31  # Column AF
for row in range(2, ws.max_row+1):
    cell = ws.cell(row=row, column=rr_col)
    if cell.value >= 2:
        cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    elif 1 <= cell.value < 2:
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    else:
        cell.fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

wb.save(output_file)
print(f"✅ Fully formatted dashboard Excel saved to {output_file}")
